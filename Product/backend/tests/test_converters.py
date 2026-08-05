"""Tests for converters."""

import csv
import io
import json
import tempfile
from pathlib import Path

from backend.converters.json_to_dmn import convert_to_dmn
from backend.converters.json_to_mermaid import convert_to_mermaid
from backend.converters.json_to_csv import convert_to_csv, convert_predicates_to_csv, convert_phrases_to_csv
from backend.converters.json_to_xlsx import convert_to_xlsx

FIXTURES = Path(__file__).parent / "fixtures"


def load_fixture(name: str) -> dict:
    with open(FIXTURES / name) as f:
        return json.load(f)


class TestDMNConverter:
    def test_produces_valid_xml(self):
        logic = load_fixture("valid_logic.json")
        xml = convert_to_dmn(logic)
        assert "<?xml" in xml
        assert "definitions" in xml

    def test_contains_decision_tables(self):
        logic = load_fixture("valid_logic.json")
        xml = convert_to_dmn(logic)
        assert "decisionTable" in xml
        assert "hitPolicy" in xml

    def test_contains_module_decisions(self):
        logic = load_fixture("valid_logic.json")
        xml = convert_to_dmn(logic)
        assert "mod_cough" in xml

    def test_contains_provenance(self):
        logic = load_fixture("valid_logic.json")
        xml = convert_to_dmn(logic)
        assert "p.38" in xml or "p.22" in xml

    def test_legacy_activator_uses_collect(self):
        logic = load_fixture("valid_logic.json")
        # The converter accepts either the current router or the legacy
        # activator. Exercise the activator contract without allowing the
        # current router to take precedence.
        logic.pop("router")
        xml = convert_to_dmn(logic)
        assert "COLLECT" in xml

    def test_current_router_uses_explicit_first_policy(self):
        logic = load_fixture("valid_logic.json")
        xml = convert_to_dmn(logic)
        assert 'hitPolicy="FIRST"' in xml


class TestMermaidConverter:
    def test_produces_flowchart(self):
        logic = load_fixture("valid_logic.json")
        mermaid = convert_to_mermaid(logic)
        assert "graph TD" in mermaid

    def test_contains_modules(self):
        logic = load_fixture("valid_logic.json")
        mermaid = convert_to_mermaid(logic)
        assert "mod_cough" in mermaid

    def test_contains_emergency_path(self):
        logic = load_fixture("valid_logic.json")
        mermaid = convert_to_mermaid(logic)
        assert "emergency" in mermaid.lower()


class TestCSVConverter:
    def test_predicates_csv_has_headers(self):
        logic = load_fixture("valid_logic.json")
        csv_str = convert_predicates_to_csv(logic)
        assert "predicate_id" in csv_str
        assert "threshold_expression" in csv_str

    def test_predicates_csv_has_data(self):
        logic = load_fixture("valid_logic.json")
        csv_str = convert_predicates_to_csv(logic)
        assert "p_danger_sign_present" in csv_str
        assert "p_fast_breathing" in csv_str

    def test_phrases_csv_has_headers(self):
        logic = load_fixture("valid_logic.json")
        csv_str = convert_phrases_to_csv(logic)
        rows = list(csv.DictReader(io.StringIO(csv_str)))
        assert rows
        assert list(rows[0]) == [
            "message_id",
            "category",
            "text",
            "module_context",
            "source_section_id",
        ]
        # Legacy english_text remains an accepted input, but the canonical
        # export header is text.
        assert rows[0]["text"] == logic["phrase_bank"][0]["english_text"]

    def test_phrases_csv_has_data(self):
        logic = load_fixture("valid_logic.json")
        csv_str = convert_phrases_to_csv(logic)
        assert "m_dx_pneumonia" in csv_str

    def test_convert_to_csv_returns_both(self):
        logic = load_fixture("valid_logic.json")
        result = convert_to_csv(logic)
        assert "predicates" in result
        assert "phrases" in result


class TestXLSXConverter:
    def test_creates_file(self):
        logic = load_fixture("valid_logic.json")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = convert_to_xlsx(logic, f.name)
            assert Path(path).exists()
            assert Path(path).stat().st_size > 0

    def test_has_required_sheets(self):
        from openpyxl import load_workbook
        logic = load_fixture("valid_logic.json")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            convert_to_xlsx(logic, f.name)
            wb = load_workbook(f.name)
            assert "survey" in wb.sheetnames
            assert "choices" in wb.sheetnames
            assert "settings" in wb.sheetnames
